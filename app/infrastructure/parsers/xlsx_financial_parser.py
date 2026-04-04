"""Parser de archivos XLSX de reportes financieros.

Extrae datos tabulares de XLSX de earnings release (como los de SQM).
Diseñado para el formato Q4Web/Nasdaq IR que usa:
- Años como integers en una fila de headers (2025, 2024)
- Contexto de período en filas merged arriba ("For the 4th quarter", "twelve months")
- Columnas de datos intercaladas con columnas vacías
- Labels en columna B (o A), valores en columnas C, E, G, I...

Sheets típicos:
- "Income Statement" → Estado de Resultados
- "Financial Statement" → Balance General (Statement of Financial Position)
- "EBITDA" → Datos complementarios (depreciation, taxes)
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)

# Keywords para detectar sheets por tipo de EEFF
_INCOME_KEYWORDS = [
    "income", "resultado", "profit", "loss", "revenue",
    "estado de resultado", "p&l",
]
_BALANCE_KEYWORDS = [
    "financial position", "financial statement", "balance",
    "posicion financiera", "assets",
]
_CASHFLOW_KEYWORDS = [
    "cash flow", "flujo", "cashflow", "cash flows",
    "statement of cash flows",
]
_EBITDA_KEYWORDS = [
    "ebitda",
]


@dataclass
class ParsedFinancialData:
    """Datos financieros extraídos de un XLSX.

    Estructura: {periodo: {metrica: valor}}
    Ejemplo:
    {
        "Q4 2025": {"Revenues": 1323.9, "Net Income": 232.9, ...},
        "FY 2025": {"Revenues": 4576.2, "Net Income": 640.6, ...},
    }
    """
    source_file: str
    ticker: str
    income_statement: dict[str, dict[str, float | None]] = field(default_factory=dict)
    balance_sheet: dict[str, dict[str, float | None]] = field(default_factory=dict)
    cash_flow: dict[str, dict[str, float | None]] = field(default_factory=dict)
    ebitda_data: dict[str, dict[str, float | None]] = field(default_factory=dict)
    currency: str = "USD"
    unit: str = "millions"
    periods_found: list[str] = field(default_factory=list)
    parsing_warnings: list[str] = field(default_factory=list)

    @property
    def has_income_statement(self) -> bool:
        return bool(self.income_statement)

    @property
    def has_balance_sheet(self) -> bool:
        return bool(self.balance_sheet)

    @property
    def has_cash_flow(self) -> bool:
        return bool(self.cash_flow)

    @property
    def completeness_score(self) -> float:
        """Score de 0 a 1 indicando completitud de los datos."""
        parts = [self.has_income_statement, self.has_balance_sheet, self.has_cash_flow]
        return sum(parts) / len(parts)


# ------------------------------------------------------------------
# Columna detectada con su índice y período asignado
# ------------------------------------------------------------------
@dataclass
class _DetectedColumn:
    """Una columna de datos detectada en el XLSX."""
    col_index: int        # Índice en la tupla de la fila
    year: int             # Año (2025, 2024...)
    period_context: str   # "quarter", "twelve months", "as of", etc.
    period_key: str       # Clave normalizada: "Q4 2025", "FY 2025", "Dec 2025"


class XLSXFinancialParser:
    """Parser de archivos XLSX de reportes financieros.

    Usa openpyxl para lectura con lógica especializada para el formato
    Q4Web/Nasdaq IR usado por SQM y empresas similares.
    """

    def parse(self, file_path: Path, ticker: str = "") -> ParsedFinancialData:
        """Parsea un XLSX y extrae los estados financieros."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl requerido para parsear XLSX. "
                "Instalar: pip install openpyxl"
            )

        logger.info(f"Parseando XLSX: {file_path.name}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames

        logger.info(f"  Sheets encontrados: {sheet_names}")

        result = ParsedFinancialData(source_file=str(file_path), ticker=ticker)

        for sheet_name in sheet_names:
            sheet_lower = sheet_name.lower()
            ws = wb[sheet_name]

            if self._matches_keywords(sheet_lower, _INCOME_KEYWORDS):
                logger.info(f"  Income Statement detectado: '{sheet_name}'")
                data, periods = self._extract_sheet_data(ws)
                if data:
                    result.income_statement = data
                    result.periods_found = list(set(result.periods_found + periods))

            elif self._matches_keywords(sheet_lower, _BALANCE_KEYWORDS):
                logger.info(f"  Balance Sheet detectado: '{sheet_name}'")
                data, periods = self._extract_sheet_data(ws)
                if data:
                    result.balance_sheet = data
                    result.periods_found = list(set(result.periods_found + periods))

            elif self._matches_keywords(sheet_lower, _CASHFLOW_KEYWORDS):
                logger.info(f"  Cash Flow detectado: '{sheet_name}'")
                data, periods = self._extract_sheet_data(ws)
                if data:
                    result.cash_flow = data
                    result.periods_found = list(set(result.periods_found + periods))

            elif self._matches_keywords(sheet_lower, _EBITDA_KEYWORDS):
                logger.info(f"  EBITDA data detectado: '{sheet_name}'")
                data, periods = self._extract_sheet_data(ws)
                if data:
                    result.ebitda_data = data
                    result.periods_found = list(set(result.periods_found + periods))

        # Detectar moneda y unidad
        self._detect_currency_unit(wb, result)
        wb.close()

        # Ordenar períodos (FY al final)
        result.periods_found.sort(key=lambda p: (
            "ZZ" if "FY" in p else p,  # FY al final
            p,
        ))

        logger.info(
            f"  Resultado: IS={result.has_income_statement}, "
            f"BS={result.has_balance_sheet}, CF={result.has_cash_flow}, "
            f"EBITDA={bool(result.ebitda_data)}, "
            f"Períodos={result.periods_found}"
        )

        return result

    # ------------------------------------------------------------------
    # Extracción de datos de un sheet
    # ------------------------------------------------------------------

    def _extract_sheet_data(
        self, ws,
    ) -> tuple[dict[str, dict[str, float | None]], list[str]]:
        """Extrae datos de un sheet individual.

        Proceso:
        1. Lee todas las filas como lista de tuplas
        2. Detecta columnas de datos (las que contienen años como integers)
        3. Construye period keys combinando contexto (filas superiores) + año
        4. Extrae label→valor para cada columna detectada
        """
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}, []

        # Paso 1: Detectar las columnas que contienen datos numéricos (años)
        columns = self._detect_data_columns(rows)
        if not columns:
            logger.warning("  No se detectaron columnas de datos con años")
            return {}, []

        periods = [c.period_key for c in columns]
        logger.info(f"  Columnas detectadas: {[(c.col_index, c.period_key) for c in columns]}")

        # Paso 2: Encontrar la fila donde empiezan los datos (después del header de años)
        data_start = self._find_data_start_row(rows, columns)

        # Paso 3: Detectar la columna de labels
        label_col = self._find_label_column(rows, data_start)

        # Paso 4: Extraer datos
        data: dict[str, dict[str, float | None]] = {p: {} for p in periods}

        for row_idx in range(data_start, len(rows)):
            row = rows[row_idx]
            if not row or label_col >= len(row) or row[label_col] is None:
                continue

            label = str(row[label_col]).strip()
            if not label:
                continue

            label = self._clean_label(label)
            if not label:
                continue

            for col in columns:
                if col.col_index < len(row):
                    raw_value = row[col.col_index]
                    parsed = self._parse_number(raw_value)
                    if parsed is not None:
                        data[col.period_key][label] = parsed

        # Limpiar períodos vacíos
        data = {p: v for p, v in data.items() if v}
        periods = list(data.keys())

        return data, periods

    def _detect_data_columns(self, rows: list[tuple]) -> list[_DetectedColumn]:
        """Detecta columnas con datos financieros buscando años como integers.

        Busca en las primeras 10 filas una fila que tenga al menos 2 celdas
        con valores de año (2015-2030). Luego lee el contexto de las filas
        anteriores para determinar el período (quarter vs full year).
        """
        year_row_idx = None
        year_cols: list[tuple[int, int]] = []  # (col_index, year)

        for idx, row in enumerate(rows[:10]):
            if not row:
                continue

            found_years = []
            for col_idx, cell in enumerate(row):
                if isinstance(cell, (int, float)) and 2015 <= cell <= 2030:
                    found_years.append((col_idx, int(cell)))

            if len(found_years) >= 2:
                year_row_idx = idx
                year_cols = found_years
                break

        if year_row_idx is None:
            return []

        # Ahora buscar contexto de período en las filas superiores
        context_map = self._build_period_context(rows, year_row_idx, year_cols)

        columns = []
        for col_idx, year in year_cols:
            context = context_map.get(col_idx, "")
            period_key = self._build_period_key(context, year)
            columns.append(_DetectedColumn(
                col_index=col_idx,
                year=year,
                period_context=context,
                period_key=period_key,
            ))

        return columns

    def _build_period_context(
        self,
        rows: list[tuple],
        year_row_idx: int,
        year_cols: list[tuple[int, int]],
    ) -> dict[int, str]:
        """Construye un mapa col_index → contexto de período.

        Lee las filas superiores al año para encontrar texto como:
        "For the 4th quarter", "For the twelve months ended December 31",
        "As of Dec. 31", etc.

        Solo considera textos que contienen keywords de período (quarter,
        month, year, "as of", etc.) para evitar confusión con títulos.

        Maneja merged cells: si una celda de contexto cubre varias columnas,
        asigna ese contexto a todas las columnas de datos dentro del rango.
        """
        _PERIOD_CONTEXT_KEYWORDS = [
            "quarter", "month", "year", "annual", "as of", "al ",
            "trimestre", "meses", "anual",
            "1st", "2nd", "3rd", "4th", "first", "second", "third", "fourth",
        ]

        context_map: dict[int, str] = {}

        # Buscar en filas anteriores al año (más cercana primero para prioridad)
        for row_idx in range(year_row_idx - 1, max(-1, year_row_idx - 4), -1):
            if row_idx < 0:
                break
            row = rows[row_idx]
            if not row:
                continue

            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                text = str(cell).strip().lower()
                if not text or len(text) < 5:
                    continue

                # Solo considerar textos con keywords de período
                if not any(kw in text for kw in _PERIOD_CONTEXT_KEYWORDS):
                    continue

                covered_year_cols = self._find_covered_columns(
                    col_idx, text, year_cols, row
                )

                for yc_idx, _ in covered_year_cols:
                    if yc_idx not in context_map:
                        context_map[yc_idx] = text

        return context_map

    def _find_covered_columns(
        self,
        context_col: int,
        context_text: str,
        year_cols: list[tuple[int, int]],
        row: tuple,
    ) -> list[tuple[int, int]]:
        """Determina qué columnas de año cubre un texto de contexto.

        Para merged cells, el texto aparece en la primera columna del merge.
        Las columnas cubiertas son las que están entre esta posición y la
        siguiente celda de contexto (o el final de la fila).
        """
        if not year_cols:
            return []

        # Encontrar la siguiente celda no-vacía después del contexto
        next_context_col = len(row)
        for col_idx, cell in enumerate(row):
            if col_idx > context_col and cell is not None:
                text = str(cell).strip()
                if text and len(text) > 3:
                    next_context_col = col_idx
                    break

        # Las columnas de datos cubiertas están entre context_col y next_context_col
        covered = [
            (col_idx, year)
            for col_idx, year in year_cols
            if context_col <= col_idx < next_context_col
        ]

        # Si no cubrió ninguna, asignar a la columna más cercana a la derecha
        if not covered:
            closest = min(
                year_cols,
                key=lambda yc: abs(yc[0] - context_col),
            )
            covered = [closest]

        return covered

    def _build_period_key(self, context: str, year: int) -> str:
        """Construye la clave de período a partir del contexto y año.

        Ejemplos:
        - "for the 4th quarter" + 2025 → "Q4 2025"
        - "for the twelve months ended december 31" + 2025 → "FY 2025"
        - "as of dec. 31" + 2025 → "Dec 2025"
        - "" + 2025 → "2025" (fallback)
        """
        ctx = context.lower()

        # Quarter detection
        quarter_patterns = [
            (r'(?:1st|first)\s*quarter', "Q1"),
            (r'(?:2nd|second)\s*quarter', "Q2"),
            (r'(?:3rd|third)\s*quarter', "Q3"),
            (r'(?:4th|fourth)\s*quarter', "Q4"),
            (r'\bq1\b', "Q1"),
            (r'\bq2\b', "Q2"),
            (r'\bq3\b', "Q3"),
            (r'\bq4\b', "Q4"),
            (r'\b1q\b', "Q1"),
            (r'\b2q\b', "Q2"),
            (r'\b3q\b', "Q3"),
            (r'\b4q\b', "Q4"),
        ]
        for pattern, quarter in quarter_patterns:
            if re.search(pattern, ctx):
                return f"{quarter} {year}"

        # Full year / annual detection
        if any(kw in ctx for kw in [
            "twelve months", "full year", "annual", "12 month",
            "doce meses", "año completo",
        ]):
            return f"FY {year}"

        # Balance sheet date
        if "as of" in ctx or "al " in ctx:
            # Buscar mes
            month_match = re.search(
                r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)',
                ctx,
            )
            if month_match:
                month_map = {
                    "jan": "Mar", "feb": "Feb", "mar": "Mar",
                    "apr": "Apr", "may": "May", "jun": "Jun",
                    "jul": "Jul", "aug": "Aug", "sep": "Sep",
                    "oct": "Oct", "nov": "Nov", "dec": "Dec",
                }
                month = month_map.get(month_match.group(1), "Dec")
                return f"{month} {year}"
            return f"Dec {year}"

        # Fallback: solo año
        return str(year)

    def _find_data_start_row(
        self,
        rows: list[tuple],
        columns: list[_DetectedColumn],
    ) -> int:
        """Encuentra la primera fila con datos numéricos después de los headers."""
        if not columns:
            return 0

        year_row = None
        for idx, row in enumerate(rows[:10]):
            if not row:
                continue
            for col in columns:
                if col.col_index < len(row):
                    cell = row[col.col_index]
                    if isinstance(cell, (int, float)) and 2015 <= cell <= 2030:
                        year_row = idx
                        break
            if year_row is not None:
                break

        if year_row is None:
            return 0

        # Los datos empiezan después de la fila de años (+ posibles filas en blanco)
        for idx in range(year_row + 1, min(year_row + 5, len(rows))):
            row = rows[idx]
            if not row:
                continue
            # Verificar si hay al menos un valor numérico en columnas de datos
            for col in columns:
                if col.col_index < len(row):
                    cell = row[col.col_index]
                    if isinstance(cell, (int, float)) and not (2015 <= cell <= 2030):
                        return idx
            # O si hay un label en la columna de labels
            for cell in row:
                if cell is not None and isinstance(cell, str) and len(cell.strip()) > 2:
                    return idx

        return year_row + 2  # Fallback

    def _find_label_column(self, rows: list[tuple], data_start: int) -> int:
        """Encuentra la columna que contiene los labels de las métricas.

        Busca en las primeras filas de datos la columna con más strings no vacíos.
        """
        if data_start >= len(rows):
            return 0

        # Contar strings por columna en las primeras filas de datos
        col_counts: dict[int, int] = {}
        for row_idx in range(data_start, min(data_start + 10, len(rows))):
            row = rows[row_idx]
            if not row:
                continue
            for col_idx, cell in enumerate(row):
                if cell is not None and isinstance(cell, str) and len(cell.strip()) > 2:
                    col_counts[col_idx] = col_counts.get(col_idx, 0) + 1

        if not col_counts:
            return 0

        # La columna con más strings es la de labels
        return max(col_counts, key=col_counts.get)

    # ------------------------------------------------------------------
    # Utilidades
    # ------------------------------------------------------------------

    def _parse_number(self, value) -> float | None:
        """Convierte un valor de celda a float."""
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip()

        if not text or text in ("—", "–", "-", "n/a", "N/A", "n.a.", "N.A."):
            return 0.0

        # Paréntesis = negativo
        is_negative = text.startswith("(") and text.endswith(")")
        if is_negative:
            text = text[1:-1]

        text = text.replace(",", "").replace(" ", "").replace("$", "").replace("%", "")

        try:
            result = float(text)
            return -result if is_negative else result
        except ValueError:
            return None

    def _clean_label(self, label: str) -> str:
        """Normaliza el label de una métrica financiera."""
        label = label.strip(" \t\n\r·•-–—")
        label = re.sub(r'\s*\(\d+\)\s*$', '', label)
        label = re.sub(r'\s+', ' ', label)
        return label

    def _detect_currency_unit(self, wb, result: ParsedFinancialData):
        """Detecta moneda y unidad (millones, miles) del XLSX."""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(max_row=8, values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).lower()

                    if "usd" in text or "us$" in text or "dollar" in text:
                        result.currency = "USD"
                    elif "clp" in text or "peso" in text or "m$" in text:
                        result.currency = "CLP"

                    if "million" in text or "millones" in text:
                        result.unit = "millions"
                    elif "thousand" in text or "miles" in text:
                        result.unit = "thousands"
                    elif "billion" in text:
                        result.unit = "billions"

    @staticmethod
    def _matches_keywords(text: str, keywords: list[str]) -> bool:
        """Verifica si el texto contiene alguno de los keywords."""
        return any(kw in text for kw in keywords)
