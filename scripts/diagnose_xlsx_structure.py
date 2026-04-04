"""Diagnóstico de la estructura real del XLSX de SQM.

Ejecutar después de haber descargado el XLSX con test_ir_scraper.py:
    python scripts/diagnose_xlsx_structure.py

Muestra las primeras filas de cada hoja para entender el formato real.
"""
import sys
from pathlib import Path

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import openpyxl

# Buscar el XLSX descargado
xlsx_paths = list(Path("data/reports").rglob("*.xlsx"))
if not xlsx_paths:
    print("No se encontraron archivos XLSX en data/reports/")
    print("Ejecuta primero: python scripts/test_ir_scraper.py")
    sys.exit(1)

xlsx_path = xlsx_paths[0]
print(f"Analizando: {xlsx_path}")
print(f"Tamaño: {xlsx_path.stat().st_size:,} bytes")
print("=" * 80)

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: '{sheet_name}' (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*80}")

    # Mostrar las primeras 25 filas con todos sus valores
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(25, ws.max_row), values_only=False), 1):
        values = []
        for cell in row:
            v = cell.value
            if v is not None:
                # Mostrar tipo para entender el formato
                if isinstance(v, (int, float)):
                    values.append(f"{v}")
                else:
                    values.append(f"'{str(v).strip()}'")
            else:
                values.append("None")

        # Solo imprimir filas que tienen al menos un valor no-None
        non_none = [v for v in values if v != "None"]
        if non_none:
            print(f"  Row {row_idx:3d}: {' | '.join(values[:12])}")  # max 12 cols

    # Mostrar merged cells si hay
    if ws.merged_cells.ranges:
        print(f"\n  Merged cells: {list(ws.merged_cells.ranges)[:10]}")

print("\n" + "=" * 80)
print("DIAGNÓSTICO COMPLETADO")
